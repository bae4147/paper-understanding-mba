"""
Merge sessions_with_classcode.csv into merged_clean.xlsx
Output: data/26-05-06/merged_final.xlsx + data/26-05-06/codebook_final.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from copy import copy

# ── 1. Load data ──────────────────────────────────────────────────────────────

sc = pd.read_csv("data/final-dataset/sessions_with_classcode.csv")

wb = load_workbook("data/26-05-06/merged_clean.xlsx")
ws = wb["merged_clean"]
mc_rows = list(ws.iter_rows(values_only=True))
mc_headers = list(mc_rows[0])
mc_data = mc_rows[1:]

# ── 2. Cassius Hudson: keep first completed session ────────────────────────────

sc_ch = sc[sc["fullName"] == "Cassius Hudson"].copy()
sc_ch["completedAt_dt"] = pd.to_datetime(sc_ch["S1_completedAt"])
sc_ch = sc_ch.sort_values("completedAt_dt")
sc = sc[sc["fullName"] != "Cassius Hudson"]
sc = pd.concat([sc, sc_ch.head(1).drop(columns=["completedAt_dt"])], ignore_index=True)

# ── 3. Build merge key on sessions side ───────────────────────────────────────

# Manual mapping: sessions fullName (exact) → merged_clean name_key
MANUAL_MAP = {
    "Bradley Dsa":               "bradley d'sa",
    "Daniel Rice":               "dan rice",
    "deffenbaugh.10@osu.edu":    "chelsea deffenbaugh",
    "esler.20@osu.edu":          "abbey esler",
    "Janette Elizabeth Stalnaker": "janette stalnaker",
    "Joe Lotozo":                "joseph lotozo",
    "Joseph Sexton":             "joe sexton",
    "karynfrost@gmail.com":      "karyn frost",
    "Kelset Gable":              "kelsey gable",
    "laxmi.mehta@osumc.edu":     "laxmi mehta",
    "Rudolph.83@osu.edu":        "tommy rudolph",
    "S Faiz Jafri":              "faiz jafri",
    "Zachary Stephens":          "zach stephens",
}

def make_merge_key(name):
    if name in MANUAL_MAP:
        return MANUAL_MAP[name]
    return name.lower()

sc["_merge_key"] = sc["fullName"].apply(make_merge_key)

# ── 4. Verify 1:1 ─────────────────────────────────────────────────────────────

dup_keys = sc[sc.duplicated("_merge_key", keep=False)]
if not dup_keys.empty:
    print("⚠️  sessions에 중복 merge_key 존재:")
    print(dup_keys[["fullName", "_merge_key"]])

# ── 5. Columns to add from sessions (explicit order) ──────────────────────────

EXCLUDE = {"fullName", "duplicate_account", "currentPhase", "surveyCompletedAt", "S3_postTask_completedAt"}

SESSION_COLS = [
    # S1 — User & session info
    "S1_userid", "S1_ClassCode", "S1_email", "S1_class", "S1_condition", "S1_userCreatedAt",
    "S1_sessionid", "S1_startedAt", "S1_completedAt",
    "S1_paperId", "S1_paper_title", "S1_paper_firstAuthorLastName", "S1_paper_year", "S1_paper_venue", "S1_paper_numPages",
    # S2 — Reading behavior
    "S2_reading_totalDuration_ms",
    "S2_window_active_ms", "S2_window_inactive_ms",
    "S2_focusTime_reading_ms", "S2_focusTime_chat_ms", "S2_focusTime_audio_ms",
    "S2_focusTime_infographics_ms", "S2_focusTime_video_ms",
    "S2_resourcesUsed_infographics", "S2_resourcesUsed_video", "S2_resourcesUsed_audio",
    "S2_resourcesUsed_audioInteractions", "S2_resourcesUsed_videoInteractions", "S2_resourcesUsed_tabSwitchCount",
    "S2_window_active_reading_ms", "S2_window_active_chat_ms", "S2_window_active_audio_ms",
    "S2_window_active_video_ms", "S2_window_active_infographics_ms",
    "S2_audio_playback_duration_sec", "S2_video_playback_duration_sec",
    "S2_chat_count", "S2_chat_history_json",
    # S3 — Post-task & survey
    "S3_postTask_contextTranslation_json", "S3_postTask_strategies_json",
    "S3_postTask_newStrategyConfidence", "S3_postTask_implementationLikelihood",
    "S3_nasa_mentalDemand", "S3_nasa_physicalDemand", "S3_nasa_temporalDemand",
    "S3_nasa_performance", "S3_nasa_effort", "S3_nasa_frustration",
    "S3_se_oc_overallGoal", "S3_se_oc_authorsReasoning", "S3_se_oc_connectingIdeas",
    "S3_se_oc_evidenceUnderstanding", "S3_se_oc_keyResultsUnderstanding", "S3_se_oc_keyTermsUnderstanding",
    "S3_se_oc_researchDesignUnderstanding", "S3_se_oc_sampleContextUnderstanding", "S3_se_oc_limitationsUnderstanding",
    "S3_se_ce_ownIdeas", "S3_se_ce_alternativePerspectives", "S3_se_ce_verifyCredibility",
    "S3_se_ce_questionClaims", "S3_se_ce_broaderImplications",
    "S3_se_ap_contextDifferences", "S3_se_ap_differencesImpact", "S3_se_ap_mechanisms",
    "S3_se_ap_outcomes", "S3_se_ap_confidence", "S3_se_ap_decision",
    "S3_attn_focus", "S3_attn_stronglyDisagreeCheck",
    "S3_llmUsefulness_overall", "S3_llmUsefulness_conceptHelp", "S3_llmUsefulness_findingsHelp",
    "S3_llmUsefulness_practicalHelp", "S3_llmUsefulness_timeSaving",
    "S3_llmTrust_competence", "S3_llmTrust_accuracy", "S3_llmTrust_benevolence", "S3_llmTrust_reliability",
    "S3_llmTrust_comfortActing", "S3_llmTrust_comfortUsing", "S3_llmTrust_relyWithoutReading",
    "S3_llmTrust_assumeClearIsAccurate", "S3_llmTrust_confidentWithoutDetails", "S3_llmTrust_relyForImportance",
    "S3_resourceHelp_chatbot", "S3_resourceHelp_infographic", "S3_resourceHelp_video", "S3_resourceHelp_podcast",
    # S4 — External tools & feedback
    "S4_externalAi_used", "S4_externalAi_details", "S4_externalPdf_used", "S4_externalPdf_details",
    "S4_studyFeedback",
]

# Verify no columns dropped or added
all_sc_cols = set(c for c in sc.columns if c not in EXCLUDE and c != "_merge_key")
assert set(SESSION_COLS) == all_sc_cols, f"Column mismatch: {all_sc_cols.symmetric_difference(set(SESSION_COLS))}"

# ── 6. Build lookup dict: merge_key → session row ─────────────────────────────

sc_indexed = sc.set_index("_merge_key")

# ── 7. Build merged rows ───────────────────────────────────────────────────────

new_headers = mc_headers + SESSION_COLS
new_rows = []

matched = 0
unmatched = 0

for row in mc_data:
    name_key = row[1]  # name_key is column index 1
    key = name_key.lower() if name_key else ""
    if key in sc_indexed.index:
        session_vals = tuple(sc_indexed.loc[key, SESSION_COLS])
        matched += 1
    else:
        session_vals = tuple([None] * len(SESSION_COLS))
        unmatched += 1
    new_rows.append(tuple(row) + session_vals)

print(f"✅ 매칭 성공: {matched}명")
print(f"⚪ 매칭 없음 (NaN): {unmatched}명")
print(f"총 rows: {len(new_rows)}")

# ── 8. Write merged xlsx ──────────────────────────────────────────────────────

out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "merged_final"

out_ws.append(new_headers)
for row in new_rows:
    out_ws.append(list(row))

out_wb.save("data/26-05-06/merged_final.xlsx")
print("✅ data/26-05-06/merged_final.xlsx 저장 완료")

# ── 9. Build codebook ─────────────────────────────────────────────────────────

# Load existing codebook
cb_wb = load_workbook("data/26-05-06/codebook.xlsx")
cb_ws = cb_wb["Sheet1"]
existing_vars = set(row[0] for row in cb_ws.iter_rows(values_only=True) if row[0])

# Field label definitions for new session fields
SESSION_LABELS = {
    # S1 — User / session metadata
    "S1_userid":               "Firebase user ID",
    "S1_ClassCode":            "Class section numeric code (from roster)",
    "S1_email":                "User email address",
    "S1_class":                "Course section (e.g., mba6202_tue)",
    "S1_condition":            "Experimental condition (control / ebm_cimo)",
    "S1_userCreatedAt":        "Account creation datetime (UTC, ISO)",
    "S1_sessionid":            "Session unique ID",
    "S1_startedAt":            "Session start datetime (UTC, ISO)",
    "S1_completedAt":          "Session completion datetime (UTC, ISO)",
    "S1_paperId":              "Assigned paper ID",
    "S1_paper_title":          "Paper title",
    "S1_paper_firstAuthorLastName": "First author last name",
    "S1_paper_year":           "Publication year",
    "S1_paper_venue":          "Journal / venue name",
    "S1_paper_numPages":       "Number of pages",
    # S2 — Reading behavior
    "S2_reading_totalDuration_ms": "Total reading phase duration (ms)",
    "S2_window_active_ms":   "Total time browser window was active (visible/focused tab) during reading phase (ms). For sessions with no window state changes (79/143), window was always active so equals reading_totalDuration_ms. Derived from window_focus_change events in reading_events.csv, clamped to reading phase boundaries. window_active_ms + window_inactive_ms = reading_totalDuration_ms.",
    "S2_window_inactive_ms": "Total time browser window was inactive (tab hidden or unfocused) during reading phase (ms). window_active_ms + window_inactive_ms = reading_totalDuration_ms.",
    "S2_focusTime_reading_ms": "Time focused on PDF reading panel during reading phase (ms). Recomputed from focus_switch events clamped to [reading_startedAt, reading_completedAt]. sum(focusTime_*) = reading_totalDuration_ms.",
    "S2_focusTime_chat_ms":    "Time focused on chatbot panel during reading phase (ms). Recomputed from focus_switch events clamped to reading phase.",
    "S2_focusTime_audio_ms":   "Time focused on audio/podcast panel during reading phase (ms). Recomputed from focus_switch events clamped to reading phase.",
    "S2_focusTime_infographics_ms": "Time focused on infographics panel during reading phase (ms). Recomputed from focus_switch events clamped to reading phase.",
    "S2_focusTime_video_ms":   "Time focused on video panel during reading phase (ms). Recomputed from focus_switch events clamped to reading phase.",
    "S2_resourcesUsed_infographics": "Whether infographics tab was opened (bool)",
    "S2_resourcesUsed_video":  "Whether video tab was opened (bool)",
    "S2_resourcesUsed_audio":  "Whether audio tab was opened (bool)",
    "S2_resourcesUsed_audioInteractions": "Audio interaction event count (play+pause+seek+end)",
    "S2_resourcesUsed_videoInteractions": "Video interaction event count (play+pause+seek+end)",
    "S2_resourcesUsed_tabSwitchCount": "Resource tab switch count",
    "S2_window_active_reading_ms":      "Time focused on reading panel while browser window was active (ms). Derived from focus_switch × window_focus_change events. sum(window_active_*_ms) = window_active_ms (within ~100ms rounding).",
    "S2_window_active_chat_ms":         "Time focused on chat panel while browser window was active (ms). Derived from focus_switch × window_focus_change events.",
    "S2_window_active_audio_ms":        "Time focused on audio panel while browser window was active (ms). Derived from focus_switch × window_focus_change events.",
    "S2_window_active_video_ms":        "Time focused on video panel while browser window was active (ms). Derived from focus_switch × window_focus_change events.",
    "S2_window_active_infographics_ms": "Time focused on infographics panel while browser window was active (ms). Derived from focus_switch × window_focus_change events.",
    "S2_audio_playback_duration_sec": "Actual audio playback duration (sec). Sum of (pause/ended currentTime − play currentTime) across all play→pause/ended pairs; seeked during playback closes the current segment and resumes from the new position.",
    "S2_video_playback_duration_sec": "Actual video playback duration (sec). Sum of (pause/ended currentTime − play currentTime) across all play→pause/ended pairs; seeked during playback closes the current segment and resumes from the new position.",
    "S2_chat_count":           "Number of chatbot queries",
    "S2_chat_history_json":    "Full chat history JSON ([{question, answer, responseTime}])",
    # S3 — Post-task (from post-task.html)
    "S3_postTask_contextTranslation_json":  "Post-task Q1 - Considering how to translate the article's conclusions to your own context, what are the key differences between the research context and your own situation? Would these matter for applying the findings? (open text, JSON array)",
    "S3_postTask_strategies_json":          "Post-task Q2 - What are practical takeaways from this article that you could realistically implement or test in your workplace? Briefly explain your approach. (open text, JSON array)",
    "S3_postTask_newStrategyConfidence":    "Post-task Q3 - How confident are you that your takeaways will be effective in your workplace? (1–7, Not at all confident–Extremely confident)",
    "S3_postTask_implementationLikelihood": "Post-task Q4 - How likely are you to actually implement these takeaways? (1–7, Very unlikely–Very likely)",
    # NASA-TLX (1–7, Very Low–Very High) — from post-study-survey.html Section 1
    "S3_nasa_mentalDemand":    "NASA-TLX - Mental Demand: How mentally demanding was the task? (1–7, Very Low–Very High)",
    "S3_nasa_physicalDemand":  "NASA-TLX - Physical Demand: How physically demanding was the task? (1–7, Very Low–Very High)",
    "S3_nasa_temporalDemand":  "NASA-TLX - Temporal Demand: How hurried or rushed was the pace of the task? (1–7, Very Low–Very High)",
    "S3_nasa_performance":     "NASA-TLX - Performance: How successful were you in understanding the article? (1–7, Poor–Excellent)",
    "S3_nasa_effort":          "NASA-TLX - Effort: How much effort did you need to put in to read and understand the article? (1–7, Very Low–Very High)",
    "S3_nasa_frustration":     "NASA-TLX - Frustration: How insecure, discouraged, irritated, stressed, or annoyed were you during the task? (1–7, Very Low–Very High)",
    # Self-Efficacy — Overall Comprehension (1–7, Strongly disagree–Strongly agree) — Section 2
    "S3_se_oc_overallGoal":                "Self-Efficacy(Overall Comprehension) - I understood the overall goal and main takeaways of the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_authorsReasoning":           "Self-Efficacy(Overall Comprehension) - I understood the authors' explanations and reasoning presented in the main text. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_connectingIdeas":            "Self-Efficacy(Overall Comprehension) - I was able to connect different sections or ideas across the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_evidenceUnderstanding":      "Self-Efficacy(Overall Comprehension) - I understood what evidence the authors used to support their claims. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_keyResultsUnderstanding":    "Self-Efficacy(Overall Comprehension) - I understood the key results/findings reported in the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_keyTermsUnderstanding":      "Self-Efficacy(Overall Comprehension) - I understood the key concepts/terms used in the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_researchDesignUnderstanding":"Self-Efficacy(Overall Comprehension) - I understood the research design used in the article (e.g., survey, experiment, interviews). (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_sampleContextUnderstanding": "Self-Efficacy(Overall Comprehension) - I understood who or what was studied (sample/context) and why that matters. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_oc_limitationsUnderstanding":   "Self-Efficacy(Overall Comprehension) - I understood the main limitations the authors described (if any). (1–7, Strongly disagree–Strongly agree)",
    # Self-Efficacy — Critical Engagement (1–7, Strongly disagree–Strongly agree) — Section 2
    "S3_se_ce_ownIdeas":               "Self-Efficacy(Critical Engagement) - I explored my own ideas and interpretations related to the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ce_alternativePerspectives":"Self-Efficacy(Critical Engagement) - I generated alternative perspectives that challenged some of the article's ideas. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ce_verifyCredibility":      "Self-Efficacy(Critical Engagement) - I thought about how to verify the credibility of the information sources cited in the article. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ce_questionClaims":         "Self-Efficacy(Critical Engagement) - I questioned what I read to judge whether the claims were convincing. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ce_broaderImplications":    "Self-Efficacy(Critical Engagement) - I reflected on the broader implications of the article (e.g., for my team, organization, or decisions). (1–7, Strongly disagree–Strongly agree)",
    # Self-Efficacy — Applicability (1–7, Strongly disagree–Strongly agree) — Section 2
    "S3_se_ap_contextDifferences":  "Self-Efficacy(Applicability) - I can clearly identify the key differences between the research context and my own situation. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ap_differencesImpact":   "Self-Efficacy(Applicability) - I can tell which of these differences would matter for applying the findings to my situation. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ap_mechanisms":          "Self-Efficacy(Applicability) - I can judge whether the underlying mechanisms behind the study's results would also apply in my situation. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ap_outcomes":            "Self-Efficacy(Applicability) - I can assess whether I would achieve similar outcomes if I applied the article's conclusions in my situation. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ap_confidence":          "Self-Efficacy(Applicability) - I feel confident in judging the applicability of this research to my specific situation. (1–7, Strongly disagree–Strongly agree)",
    "S3_se_ap_decision":            "Self-Efficacy(Applicability) - I can make an informed decision about whether to apply these findings in my context. (1–7, Strongly disagree–Strongly agree)",
    # Attention check (1–5, Strongly disagree–Strongly agree) — Section 5 final
    "S3_attn_focus":                   "Attention - In general, I was focused when reading and responding to the survey items. (1–5, Strongly disagree–Strongly agree)",
    "S3_attn_stronglyDisagreeCheck":   "Attention - Click 'strongly disagree' if you are still paying attention. (1–5, trap item; correct = 1)",
    # LLM Usefulness (1–7) — Section 3, from post-study-survey.html
    "S3_llmUsefulness_overall":        "LLM Usefulness - How useful was the GenAI for reading and understanding the article? (1–7, Not at all useful–Extremely useful)",
    "S3_llmUsefulness_conceptHelp":    "LLM Usefulness - How helpful was the GenAI for understanding difficult concepts or terminology? (1–7, Not at all helpful–Extremely helpful)",
    "S3_llmUsefulness_findingsHelp":   "LLM Usefulness - How helpful was the GenAI for understanding the research findings? (1–7, Not at all helpful–Extremely helpful)",
    "S3_llmUsefulness_practicalHelp":  "LLM Usefulness - How helpful was the GenAI for understanding practical applications? (1–7, Not at all helpful–Extremely helpful)",
    "S3_llmUsefulness_timeSaving":     "LLM Usefulness - How helpful was the GenAI for saving your time? (1–7, Not at all helpful–Extremely helpful)",
    # LLM Trust (1–7, Strongly disagree–Strongly agree) — Section 4
    "S3_llmTrust_competence":              "LLM Trust - Overall, the GenAI was competent and effective as a tool for assisting with reading. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_accuracy":                "LLM Trust - The information provided by the GenAI was accurate. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_benevolence":             "LLM Trust - I believe the GenAI is made with my best interests in mind. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_reliability":             "LLM Trust - I can always rely on the GenAI when understanding articles. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_comfortActing":           "LLM Trust - I have no reservations about acting on the information provided by the GenAI. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_comfortUsing":            "LLM Trust - I have no hesitation in using the information provided by the GenAI. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_relyWithoutReading":      "LLM Trust - I can rely on the GenAI to understand an article even when I don't fully read it myself. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_assumeClearIsAccurate":   "LLM Trust - If the GenAI's summary seems clear, I assume it is accurate. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_confidentWithoutDetails": "LLM Trust - The GenAI's explanation makes me feel confident about an article even when I can't identify supporting details in the text. (1–7, Strongly disagree–Strongly agree)",
    "S3_llmTrust_relyForImportance":       "LLM Trust - I rely on the GenAI to tell me what matters instead of deciding for myself what is important in the article. (1–7, Strongly disagree–Strongly agree)",
    # Resource helpfulness (0–7, Did not use–Extremely helpful) — Section 5
    "S3_resourceHelp_chatbot":     "Resource Help - How helpful was the AI chatbot? (0–7, Did not use–Extremely helpful)",
    "S3_resourceHelp_infographic": "Resource Help - How helpful were the infographics? (0–7, Did not use–Extremely helpful)",
    "S3_resourceHelp_video":       "Resource Help - How helpful was the video? (0–7, Did not use–Extremely helpful)",
    "S3_resourceHelp_podcast":     "Resource Help - How helpful was the podcast? (0–7, Did not use–Extremely helpful)",
    # S4 — External tools & feedback
    "S4_externalAi_used":      "External AI tool used during task (yes/no)",
    "S4_externalAi_details":   "Details of external AI tool used (open text)",
    "S4_externalPdf_used":     "External PDF viewer used during task (yes/no)",
    "S4_externalPdf_details":  "Details of external PDF viewer used (open text)",
    "S4_studyFeedback":        "Free-text feedback about the study",
}

# Append new rows to codebook
added = 0
for col in SESSION_COLS:
    if col not in existing_vars:
        label = SESSION_LABELS.get(col, col)
        cb_ws.append([col, label])
        added += 1

cb_wb.save("data/26-05-06/codebook_final.xlsx")
print(f"✅ data/26-05-06/codebook_final.xlsx 저장 완료 ({added}개 항목 추가)")
